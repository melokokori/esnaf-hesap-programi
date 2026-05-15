import customtkinter as ctk
import sqlite3, os, sys, shutil
from tkinter import ttk, messagebox, StringVar, filedialog
from datetime import datetime

VERSION = "1.0.0"
COPYRIGHT = "© 2025 melokokori"

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH  = os.path.join(BASE_DIR, "hesap.db")
XLS_PATH = os.path.join(BASE_DIR, "..", "nevinbebe_ozet.xlsx")

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


def show_splash():
    splash = ctk.CTk()
    splash.overrideredirect(True)
    splash.configure(fg_color="#2F5496")
    w, h = 440, 220
    sw = splash.winfo_screenwidth()
    sh = splash.winfo_screenheight()
    splash.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    ctk.CTkLabel(splash, text="Esnaf Hesap Programı",
                 font=ctk.CTkFont(size=24, weight="bold"),
                 text_color="white").pack(expand=True, pady=(40, 4))
    ctk.CTkLabel(splash, text=f"Sürüm {VERSION}",
                 font=ctk.CTkFont(size=13),
                 text_color="#BDD7EE").pack()
    ctk.CTkLabel(splash, text=COPYRIGHT,
                 font=ctk.CTkFont(size=11),
                 text_color="#8FB4D4").pack(pady=(4, 40))

    splash.after(2500, splash.destroy)
    splash.mainloop()

# ── VERİTABANI ────────────────────────────────────────────────
def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS musteriler (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            musteri_adi TEXT NOT NULL,
            telefon     TEXT DEFAULT '',
            arsiv       INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS islemler (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            musteri_id  INTEGER NOT NULL REFERENCES musteriler(id),
            tarih       TEXT,
            aciklama    TEXT DEFAULT '',
            fiyat       REAL DEFAULT 0,
            odenen      REAL DEFAULT 0,
            kalan       REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS gunluk_satis (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            tarih       TEXT NOT NULL,
            isim        TEXT NOT NULL,
            urun        TEXT NOT NULL,
            marka       TEXT DEFAULT '',
            urun_turu   TEXT DEFAULT '',
            fiyat       REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS markalar (
            id  INTEGER PRIMARY KEY AUTOINCREMENT,
            adi TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS urun_turleri (
            id  INTEGER PRIMARY KEY AUTOINCREMENT,
            adi TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS config (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
    """)
    con.commit()
    return con

def get_config(con, key, default=""):
    row = con.cursor().execute("SELECT value FROM config WHERE key=?", (key,)).fetchone()
    return row[0] if row else default

def set_config(con, key, value):
    con.cursor().execute("INSERT OR REPLACE INTO config (key,value) VALUES (?,?)", (key, value))
    con.commit()

def import_from_excel(con):
    try:
        import openpyxl
    except ImportError:
        return False
    if not os.path.exists(XLS_PATH):
        return False
    wb  = openpyxl.load_workbook(XLS_PATH, read_only=True, data_only=True)
    ws_ozet   = wb["Musteri Ozeti"]
    ws_detail = wb["Tum Islemler"]
    cur = con.cursor()
    musteri_map = {}
    for row in ws_ozet.iter_rows(min_row=2, values_only=True):
        sayfa, ad, tel = str(row[0] or ""), str(row[1] or "").strip(), str(row[2] or "").strip()
        if not ad:
            continue
        cur.execute("INSERT INTO musteriler (musteri_adi, telefon) VALUES (?,?)", (ad, tel))
        musteri_map[sayfa] = cur.lastrowid
    con.commit()
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        sayfa = str(row[0] or "")
        mid   = musteri_map.get(sayfa)
        if mid is None:
            continue
        raw = row[3]
        tarih  = raw.strftime("%d.%m.%Y") if hasattr(raw, "strftime") else str(raw or "").strip()
        acikl  = str(row[4] or "").strip()
        fiyat  = float(row[5]) if row[5] else 0.0
        odenen = float(row[6]) if row[6] else 0.0
        kalan  = float(row[7]) if row[7] else 0.0
        cur.execute(
            "INSERT INTO islemler (musteri_id,tarih,aciklama,fiyat,odenen,kalan) VALUES (?,?,?,?,?,?)",
            (mid, tarih, acikl, fiyat, odenen, kalan)
        )
    con.commit()
    wb.close()
    return True

# ── ANA PENCERE ───────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self, con):
        super().__init__()
        self.con = con
        self._isletme = get_config(con, "isletme_adi", "Hesap Programı")
        self.title(f"{self._isletme} — Hesap Programı")
        self.geometry("1150x700")
        self.minsize(950, 580)
        icon_path = os.path.join(BASE_DIR, "icon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
        self._build_ui()
        self.load_customers()
        if not self._isletme or self._isletme == "Hesap Programı":
            self.after(200, self._ilk_kurulum)
        self.bind("<Control-s>", lambda _: self._sessiz_kaydet())
        self.bind("<Control-r>", lambda _: self._rapor_excel())
        self.bind("<Control-h>", lambda _: self._hakkinda())

    def _build_ui(self):
        sb = ctk.CTkFrame(self, height=22, corner_radius=0, fg_color="#DDEEFF")
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        ctk.CTkLabel(sb, text=f"{COPYRIGHT}  |  v{VERSION}",
                     font=ctk.CTkFont(size=10), text_color="#5577AA").pack(side="right", padx=14)
        ctk.CTkLabel(sb, text="Ctrl+S: Kaydet   Ctrl+R: Rapor   Ctrl+H: Hakkında",
                     font=ctk.CTkFont(size=10), text_color="#8899AA").pack(side="left", padx=14)

        self.tabs = ctk.CTkTabview(self, anchor="nw")
        self.tabs.pack(fill="both", expand=True)
        self.tabs.add("📋  Müşteri Hesapları")
        self.tabs.add("🛍  Günlük Satışlar")
        self._build_hesaplar_tab(self.tabs.tab("📋  Müşteri Hesapları"))
        self._build_satis_tab(self.tabs.tab("🛍  Günlük Satışlar"))

    # ════════════════════════════════════════════
    #  SEKME 1 — MÜŞTERİ HESAPLARI
    # ════════════════════════════════════════════
    def _build_hesaplar_tab(self, parent):
        self.left = ctk.CTkFrame(parent, width=360, corner_radius=8)
        self.left.pack(side="left", fill="y", padx=(8,4), pady=8)
        self.left.pack_propagate(False)

        ctk.CTkLabel(self.left, text="Müşteri Hesapları",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(14,4))

        sf = ctk.CTkFrame(self.left, fg_color="transparent")
        sf.pack(fill="x", padx=12, pady=(0,6))
        self.search_var = StringVar()
        self.search_var.trace_add("write", lambda *_: self.load_customers())
        ctk.CTkEntry(sf, textvariable=self.search_var,
                     placeholder_text="🔍  İsim veya telefon ara...", height=34).pack(fill="x")

        bf = ctk.CTkFrame(self.left, fg_color="transparent")
        bf.pack(fill="x", padx=12, pady=(0,8))
        self.filter_var = StringVar(value="Borçlular")
        ctk.CTkSegmentedButton(bf,
            values=["Borçlular", "Tümü", "Arşiv"],
            variable=self.filter_var,
            command=lambda _: self.load_customers()
        ).pack(fill="x")

        lf = ctk.CTkFrame(self.left, fg_color="transparent")
        lf.pack(fill="both", expand=True, padx=12, pady=4)

        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=28, font=("Segoe UI", 10))
        style.configure("Custom.Treeview.Heading", font=("Segoe UI", 10, "bold"))

        self.tree = ttk.Treeview(lf, columns=("ad","kalan"),
                                  show="headings", style="Custom.Treeview", selectmode="browse")
        self.tree.heading("ad",    text="Müşteri Adı")
        self.tree.heading("kalan", text="Bakiye")
        self.tree.column("ad",    width=210, anchor="w")
        self.tree.column("kalan", width=100, anchor="e")
        self.tree.tag_configure("borclu",  foreground="#C00000")
        self.tree.tag_configure("borcsuz", foreground="#217346")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        sb = ttk.Scrollbar(lf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        ctk.CTkButton(self.left, text="+ Yeni Müşteri", height=36,
                      command=self.new_customer).pack(padx=12, pady=(8,4), fill="x")

        row1 = ctk.CTkFrame(self.left, fg_color="transparent")
        row1.pack(fill="x", padx=12, pady=(0,4))
        ctk.CTkButton(row1, text="📊 Rapor Al", height=34, width=10,
                      fg_color="#217346",
                      command=self._rapor_excel).pack(side="left", expand=True, fill="x", padx=(0,3))
        ctk.CTkButton(row1, text="💾 Yedek Al", height=34, width=10,
                      fg_color="#ED7D31",
                      command=self._yedek_al).pack(side="left", expand=True, fill="x", padx=(3,0))

        row2 = ctk.CTkFrame(self.left, fg_color="transparent")
        row2.pack(fill="x", padx=12, pady=(0,4))
        ctk.CTkButton(row2, text="⚙ Ayarlar", height=34, width=10,
                      fg_color="#5B5B5B",
                      command=self._ayarlar).pack(side="left", expand=True, fill="x", padx=(0,3))
        ctk.CTkButton(row2, text="ℹ Hakkında", height=34, width=10,
                      fg_color="#5B5B5B",
                      command=self._hakkinda).pack(side="left", expand=True, fill="x", padx=(3,0))

        ctk.CTkButton(self.left, text="💾  Kaydet ve Çık", height=36,
                      fg_color="#2F5496",
                      command=self._kaydet_ve_cik).pack(padx=12, pady=(0,10), fill="x")

        self.right = ctk.CTkFrame(parent, corner_radius=8, fg_color="#F5F5F5")
        self.right.pack(side="right", fill="both", expand=True, padx=(4,8), pady=8)
        self._show_placeholder()

    # ════════════════════════════════════════════
    #  SEKME 2 — GÜNLÜK SATIŞLAR
    # ════════════════════════════════════════════
    def _build_satis_tab(self, parent):
        form_card = ctk.CTkFrame(parent, corner_radius=10)
        form_card.pack(fill="x", padx=12, pady=(10,4))

        ctk.CTkLabel(form_card, text="Yeni Satış Ekle",
                     font=ctk.CTkFont(size=14, weight="bold")).grid(
                     row=0, column=0, columnspan=10, sticky="w", padx=16, pady=(10,4))

        ctk.CTkLabel(form_card, text="Tarih").grid(row=1, column=0, padx=(16,4), pady=6, sticky="w")
        self.tarih_var = StringVar(value=datetime.today().strftime("%d.%m.%Y %H:%M"))
        self.s_tarih = ctk.CTkEntry(form_card, width=140, textvariable=self.tarih_var)
        self.s_tarih.grid(row=1, column=1, padx=(0,16), pady=6)
        self._saat_guncelle()

        ctk.CTkLabel(form_card, text="Ad Soyad").grid(row=1, column=2, padx=(0,4), pady=6, sticky="w")
        self.s_isim = ctk.CTkEntry(form_card, width=200, placeholder_text="Müşteri adı...")
        self.s_isim.grid(row=1, column=3, padx=(0,16), pady=6)
        self.s_isim.bind("<KeyRelease>", self._isim_oneri)

        ctk.CTkLabel(form_card, text="Fiyat (TL)").grid(row=1, column=4, padx=(0,4), pady=6, sticky="w")
        self.s_fiyat = ctk.CTkEntry(form_card, width=120)
        self.s_fiyat.grid(row=1, column=5, padx=(0,16), pady=6)

        alt_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        alt_frame.grid(row=2, column=0, columnspan=10, padx=(16,16), pady=6, sticky="w")

        ctk.CTkLabel(alt_frame, text="Marka").pack(side="left", padx=(0,4))
        self.s_marka = ctk.CTkComboBox(alt_frame, width=190, values=self._get_markalar())
        self.s_marka.set("")
        self.s_marka.pack(side="left")
        ctk.CTkButton(alt_frame, text="⚙", width=32, height=32, fg_color="#5B5B5B",
                      command=lambda: self._liste_yonet("marka")).pack(side="left", padx=(4,20))

        ctk.CTkLabel(alt_frame, text="Ürün Türü").pack(side="left", padx=(0,4))
        self.s_urun_turu = ctk.CTkComboBox(alt_frame, width=190, values=self._get_urun_turleri())
        self.s_urun_turu.set("")
        self.s_urun_turu.pack(side="left")
        ctk.CTkButton(alt_frame, text="⚙", width=32, height=32, fg_color="#5B5B5B",
                      command=lambda: self._liste_yonet("urun_turu")).pack(side="left", padx=(4,20))

        ctk.CTkButton(alt_frame, text="+ Ekle", width=100, height=34,
                      fg_color="#217346",
                      command=self._satis_ekle).pack(side="left")

        self._oneri_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=6,
                                          border_width=1, border_color="#CCCCCC")

        filtre_bar = ctk.CTkFrame(parent, fg_color="transparent")
        filtre_bar.pack(fill="x", padx=12, pady=(4,4))
        ctk.CTkLabel(filtre_bar, text="Tarih filtresi:").pack(side="left", padx=(4,6))
        self.s_filtre = ctk.CTkSegmentedButton(filtre_bar,
            values=["Bugün", "Bu Hafta", "Bu Ay", "Tümü"],
            command=self._satis_yukle)
        self.s_filtre.set("Bugün")
        self.s_filtre.pack(side="left")
        ctk.CTkButton(filtre_bar, text="🛒  Toplu Giriş", width=130, height=30,
                      fg_color="#2F5496",
                      command=self._toplu_giris).pack(side="left", padx=(16,0))
        self.s_toplam_lbl = ctk.CTkLabel(filtre_bar, text="Toplam: 0,00 TL",
                                          font=ctk.CTkFont(size=13, weight="bold"),
                                          text_color="#2F5496")
        self.s_toplam_lbl.pack(side="right", padx=12)

        tbl_frame = ctk.CTkFrame(parent, corner_radius=8)
        tbl_frame.pack(fill="both", expand=True, padx=12, pady=(0,8))

        style = ttk.Style()
        style.configure("Satis.Treeview", rowheight=28, font=("Segoe UI", 10))
        style.configure("Satis.Treeview.Heading", font=("Segoe UI", 10, "bold"))

        self.s_tree = ttk.Treeview(tbl_frame,
            columns=("tarih","isim","marka","urun_turu","fiyat"),
            show="headings", style="Satis.Treeview", selectmode="browse")
        for c, lbl, w, anc in [
            ("tarih",    "Tarih",      130, "center"),
            ("isim",     "Ad Soyad",   190, "w"),
            ("marka",    "Marka",      155, "w"),
            ("urun_turu","Ürün Türü",  155, "w"),
            ("fiyat",    "Fiyat (TL)", 110, "e"),
        ]:
            self.s_tree.heading(c, text=lbl)
            self.s_tree.column(c, width=w, anchor=anc)

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.s_tree.yview)
        self.s_tree.configure(yscrollcommand=vsb.set)
        self.s_tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=8)
        vsb.pack(side="right", fill="y", pady=8, padx=(0,4))

        ctk.CTkButton(tbl_frame, text="🗑 Seçili Satırı Sil", width=160,
                      fg_color="#C00000",
                      command=self._satis_sil).pack(side="bottom", pady=(0,8))

        self._satis_yukle()

    def _saat_guncelle(self):
        # Sadece kullanıcı elle değiştirmediyse güncelle
        mevcut = self.tarih_var.get()
        beklenen_tarih = datetime.today().strftime("%d.%m.%Y")
        if mevcut.startswith(beklenen_tarih) or len(mevcut) < 10:
            self.tarih_var.set(datetime.today().strftime("%d.%m.%Y %H:%M"))
        self.after(10000, self._saat_guncelle)

    def _get_markalar(self):
        return [r[0] for r in self.con.cursor().execute("SELECT adi FROM markalar ORDER BY adi").fetchall()]

    def _get_urun_turleri(self):
        return [r[0] for r in self.con.cursor().execute("SELECT adi FROM urun_turleri ORDER BY adi").fetchall()]

    def _liste_yonet(self, tip):
        baslik = "Marka Yönetimi" if tip == "marka" else "Ürün Türü Yönetimi"
        tablo  = "markalar"       if tip == "marka" else "urun_turleri"
        combo  = self.s_marka     if tip == "marka" else self.s_urun_turu

        dlg = ctk.CTkToplevel(self)
        dlg.title(baslik)
        dlg.geometry("320x420")
        dlg.grab_set(); dlg.resizable(False, False)

        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=16, pady=14)
        ctk.CTkLabel(frame, text=baslik,
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(0,8))

        lb_frame = ctk.CTkFrame(frame, fg_color="transparent")
        lb_frame.pack(fill="both", expand=True)
        style = ttk.Style(); style.configure("YY.Treeview", rowheight=26, font=("Segoe UI",10))
        lb = ttk.Treeview(lb_frame, columns=("adi",), show="headings",
                          style="YY.Treeview", selectmode="browse", height=10)
        lb.heading("adi", text="Ad"); lb.column("adi", width=240, anchor="w")
        sb2 = ttk.Scrollbar(lb_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb2.set)
        lb.pack(side="left", fill="both", expand=True)
        sb2.pack(side="right", fill="y")

        def yenile():
            for i in lb.get_children(): lb.delete(i)
            for r in self.con.cursor().execute(f"SELECT id,adi FROM {tablo} ORDER BY adi"):
                lb.insert("", "end", iid=str(r[0]), values=(r[1],))
            combo.configure(values=self._get_markalar() if tip=="marka" else self._get_urun_turleri())

        yenile()

        ekle_f = ctk.CTkFrame(frame, fg_color="transparent")
        ekle_f.pack(fill="x", pady=(10,4))
        e_yeni = ctk.CTkEntry(ekle_f, width=200, placeholder_text="Yeni ad...")
        e_yeni.pack(side="left", padx=(0,6))

        def ekle():
            adi = e_yeni.get().strip()
            if not adi: return
            try:
                self.con.cursor().execute(f"INSERT INTO {tablo} (adi) VALUES (?)", (adi,))
                self.con.commit()
            except Exception:
                messagebox.showerror("Hata", f'"{adi}" zaten mevcut.', parent=dlg); return
            e_yeni.delete(0, "end")
            yenile()

        ctk.CTkButton(ekle_f, text="Ekle", width=70, height=32,
                      fg_color="#217346", command=ekle).pack(side="left")

        def sil():
            sel = lb.selection()
            if not sel: return
            adi = lb.item(sel[0])["values"][0]
            if not messagebox.askyesno("Onay", f'"{adi}" silinsin mi?', parent=dlg): return
            self.con.cursor().execute(f"DELETE FROM {tablo} WHERE id=?", (int(sel[0]),))
            self.con.commit()
            yenile()

        ctk.CTkButton(frame, text="🗑 Seçiliyi Sil", fg_color="#C00000",
                      command=sil).pack(fill="x", pady=(4,0))

    def _isim_oneri(self, event=None):
        q = self.s_isim.get().strip()
        if len(q) < 2:
            self._oneri_gizle(); return
        rows = self.con.cursor().execute(
            "SELECT musteri_adi FROM musteriler WHERE musteri_adi LIKE ? ORDER BY musteri_adi LIMIT 8",
            (f"%{q}%",)
        ).fetchall()
        if rows:
            self._oneri_goster([r[0] for r in rows])
        else:
            self._oneri_gizle()

    def _oneri_goster(self, isimler):
        self._oneri_gizle()
        x = self.s_isim.winfo_rootx() - self.winfo_rootx()
        y = self.s_isim.winfo_rooty() - self.winfo_rooty() + self.s_isim.winfo_height()
        self._oneri_frame.place(x=x, y=y, width=220)
        for isim in isimler:
            ctk.CTkButton(self._oneri_frame, text=isim, height=28,
                          fg_color="white", text_color="black",
                          hover_color="#E0EEFF", anchor="w",
                          command=lambda i=isim: self._oneri_sec(i)).pack(fill="x", padx=2, pady=1)
        self._oneri_frame.lift()

    def _oneri_gizle(self):
        for w in self._oneri_frame.winfo_children(): w.destroy()
        self._oneri_frame.place_forget()

    def _oneri_sec(self, isim):
        self.s_isim.delete(0, "end")
        self.s_isim.insert(0, isim)
        self._oneri_gizle()
        self.s_urun_turu.focus()

    def _satis_ekle(self):
        tarih     = self.s_tarih.get().strip()
        isim      = self.s_isim.get().strip()
        marka     = self.s_marka.get().strip()
        urun_turu = self.s_urun_turu.get().strip()
        try:
            fiyat = float(self.s_fiyat.get().replace(",","."))
        except ValueError:
            messagebox.showerror("Hata", "Geçerli bir fiyat girin."); return
        if not isim:
            messagebox.showerror("Hata", "Ad soyad boş olamaz."); return
        if not marka and not urun_turu:
            messagebox.showerror("Hata", "En az Marka veya Ürün Türü seçin."); return
        urun = " - ".join(filter(None, [marka, urun_turu]))
        self.con.cursor().execute(
            "INSERT INTO gunluk_satis (tarih,isim,urun,marka,urun_turu,fiyat) VALUES (?,?,?,?,?,?)",
            (tarih, isim, urun, marka, urun_turu, fiyat)
        )
        self.con.commit()
        self.s_fiyat.delete(0, "end")
        self.s_marka.set(""); self.s_urun_turu.set("")
        self._satis_yukle()

    def _satis_yukle(self, *_):
        for item in self.s_tree.get_children(): self.s_tree.delete(item)
        filtre = self.s_filtre.get()
        bugun  = datetime.today().strftime("%d.%m.%Y")
        cur    = self.con.cursor()
        if filtre == "Bugün":
            rows = cur.execute(
                "SELECT id,tarih,isim,marka,urun_turu,fiyat FROM gunluk_satis WHERE tarih LIKE ? ORDER BY id DESC",
                (f"{bugun}%",)).fetchall()
        elif filtre == "Bu Hafta":
            rows = cur.execute("""
                SELECT id,tarih,isim,marka,urun_turu,fiyat FROM gunluk_satis
                WHERE substr(tarih,7,4)||substr(tarih,4,2)||substr(tarih,1,2)
                      >= strftime('%Y%m%d','now','-6 days')
                ORDER BY id DESC""").fetchall()
        elif filtre == "Bu Ay":
            ay = datetime.today().strftime("%m.%Y")
            rows = cur.execute(
                "SELECT id,tarih,isim,marka,urun_turu,fiyat FROM gunluk_satis WHERE substr(tarih,4,7)=? ORDER BY id DESC",
                (ay,)).fetchall()
        else:
            rows = cur.execute(
                "SELECT id,tarih,isim,marka,urun_turu,fiyat FROM gunluk_satis ORDER BY id DESC").fetchall()
        toplam = 0.0
        for r in rows:
            self.s_tree.insert("", "end", iid=str(r[0]),
                               values=(r[1], r[2], r[3], r[4], f"{r[5]:,.2f}"))
            toplam += r[5]
        self.s_toplam_lbl.configure(text=f"Toplam: {toplam:,.2f} TL")

    def _toplu_giris(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Toplu Alışveriş Girişi")
        dlg.geometry("700x560")
        dlg.grab_set(); dlg.resizable(True, True)
        sepet = []

        ust = ctk.CTkFrame(dlg, corner_radius=8)
        ust.pack(fill="x", padx=14, pady=(12,6))
        ctk.CTkLabel(ust, text="Ad Soyad:").grid(row=0, column=0, padx=(12,4), pady=8, sticky="w")
        e_isim = ctk.CTkEntry(ust, width=220, placeholder_text="Müşteri adı...")
        e_isim.grid(row=0, column=1, padx=(0,20), pady=8)
        ctk.CTkLabel(ust, text="Tarih:").grid(row=0, column=2, padx=(0,4), pady=8, sticky="w")
        e_tarih = ctk.CTkEntry(ust, width=130)
        e_tarih.insert(0, datetime.today().strftime("%d.%m.%Y %H:%M"))
        e_tarih.grid(row=0, column=3, padx=(0,12), pady=8)

        orta = ctk.CTkFrame(dlg, corner_radius=8)
        orta.pack(fill="x", padx=14, pady=(0,6))
        ctk.CTkLabel(orta, text="Marka:").grid(row=0, column=0, padx=(12,4), pady=8, sticky="w")
        cb_marka = ctk.CTkComboBox(orta, width=170, values=self._get_markalar()); cb_marka.set("")
        cb_marka.grid(row=0, column=1, padx=(0,14), pady=8)
        ctk.CTkLabel(orta, text="Ürün Türü:").grid(row=0, column=2, padx=(0,4), pady=8, sticky="w")
        cb_urun = ctk.CTkComboBox(orta, width=170, values=self._get_urun_turleri()); cb_urun.set("")
        cb_urun.grid(row=0, column=3, padx=(0,14), pady=8)
        ctk.CTkLabel(orta, text="Fiyat:").grid(row=0, column=4, padx=(0,4), pady=8, sticky="w")
        e_fiyat = ctk.CTkEntry(orta, width=90)
        e_fiyat.grid(row=0, column=5, padx=(0,10), pady=8)

        tbl_frame = ctk.CTkFrame(dlg, corner_radius=8)
        tbl_frame.pack(fill="both", expand=True, padx=14, pady=(0,6))
        style = ttk.Style(); style.configure("Sepet.Treeview", rowheight=26, font=("Segoe UI",10))
        style.configure("Sepet.Treeview.Heading", font=("Segoe UI",10,"bold"))
        s_tree = ttk.Treeview(tbl_frame, columns=("sira","marka","urun_turu","fiyat"),
                               show="headings", style="Sepet.Treeview", selectmode="browse")
        for c,lbl,w,anc in [("sira","#",40,"center"),("marka","Marka",190,"w"),
                              ("urun_turu","Ürün Türü",190,"w"),("fiyat","Fiyat",120,"e")]:
            s_tree.heading(c,text=lbl); s_tree.column(c,width=w,anchor=anc)
        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=s_tree.yview)
        s_tree.configure(yscrollcommand=vsb.set)
        s_tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=8)
        vsb.pack(side="right", fill="y", pady=8, padx=(0,4))

        alt = ctk.CTkFrame(dlg, fg_color="transparent")
        alt.pack(fill="x", padx=14, pady=(0,12))
        toplam_lbl = ctk.CTkLabel(alt, text="Toplam: 0,00 TL",
                                   font=ctk.CTkFont(size=14, weight="bold"), text_color="#2F5496")
        toplam_lbl.pack(side="left", padx=4)

        def guncelle_toplam():
            toplam_lbl.configure(text=f"Toplam: {sum(u['fiyat'] for u in sepet):,.2f} TL")

        def sepete_ekle():
            marka_v = cb_marka.get().strip(); urun_v = cb_urun.get().strip()
            try: fiyat_v = float(e_fiyat.get().replace(",","."))
            except ValueError:
                messagebox.showerror("Hata", "Geçerli bir fiyat girin.", parent=dlg); return
            if not marka_v and not urun_v:
                messagebox.showerror("Hata", "Marka veya Ürün Türü seçin.", parent=dlg); return
            sepet.append({"marka": marka_v, "urun_turu": urun_v, "fiyat": fiyat_v})
            sira = len(sepet)
            s_tree.insert("", "end", iid=str(sira), values=(sira, marka_v, urun_v, f"{fiyat_v:,.2f}"))
            e_fiyat.delete(0, "end"); cb_marka.set(""); cb_urun.set("")
            e_fiyat.focus(); guncelle_toplam()

        def sepetten_sil():
            sel = s_tree.selection()
            if not sel: return
            sepet.pop(int(sel[0])-1)
            for i in s_tree.get_children(): s_tree.delete(i)
            for i,u in enumerate(sepet,1):
                s_tree.insert("","end",iid=str(i),values=(i,u["marka"],u["urun_turu"],f"{u['fiyat']:,.2f}"))
            guncelle_toplam()

        def kaydet():
            isim_v  = e_isim.get().strip(); tarih_v = e_tarih.get().strip()
            if not isim_v:
                messagebox.showerror("Hata","Müşteri adı boş olamaz.",parent=dlg); return
            if not sepet:
                messagebox.showerror("Hata","Sepet boş.",parent=dlg); return
            cur = self.con.cursor()
            for u in sepet:
                urun = " - ".join(filter(None,[u["marka"],u["urun_turu"]]))
                cur.execute(
                    "INSERT INTO gunluk_satis (tarih,isim,urun,marka,urun_turu,fiyat) VALUES (?,?,?,?,?,?)",
                    (tarih_v, isim_v, urun, u["marka"], u["urun_turu"], u["fiyat"])
                )
            self.con.commit()
            dlg.destroy(); self._satis_yukle()

        ctk.CTkButton(orta, text="+ Sepete Ekle", width=120, height=32,
                      fg_color="#217346", command=sepete_ekle).grid(row=0, column=6, padx=(0,12), pady=8)
        ctk.CTkButton(alt, text="🗑 Seçiliyi Sil", width=130, fg_color="#C00000",
                      command=sepetten_sil).pack(side="left", padx=(12,0))
        ctk.CTkButton(alt, text="✅  Tümünü Kaydet", width=160, fg_color="#2F5496",
                      command=kaydet).pack(side="right")
        e_fiyat.bind("<Return>", lambda _: sepete_ekle())

    def _satis_sil(self):
        sel = self.s_tree.selection()
        if not sel: return
        if not messagebox.askyesno("Onay", "Seçili kayıt silinsin mi?"): return
        self.con.cursor().execute("DELETE FROM gunluk_satis WHERE id=?", (int(sel[0]),))
        self.con.commit(); self._satis_yukle()

    # ════════════════════════════════════════════
    #  MÜŞTERİ HESAPLARI
    # ════════════════════════════════════════════
    def _show_placeholder(self):
        for w in self.right.winfo_children(): w.destroy()
        ctk.CTkLabel(self.right, text="← Listeden bir müşteri seçin",
                     font=ctk.CTkFont(size=14), text_color="gray").pack(expand=True)

    def _ilk_kurulum(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Hoş Geldiniz — Kurulum")
        dlg.geometry("400x200"); dlg.grab_set(); dlg.resizable(False, False)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)  # kapatma butonunu devre dışı bırak

        ctk.CTkLabel(dlg, text="Hesap Programına Hoş Geldiniz",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(24,6))
        ctk.CTkLabel(dlg, text="İşletmenizin adını girin:",
                     font=ctk.CTkFont(size=12)).pack()
        e = ctk.CTkEntry(dlg, width=260, placeholder_text="örn. Ahmet Usta Şarküteri")
        e.pack(pady=12); e.focus()

        def tamam():
            ad = e.get().strip()
            if not ad:
                messagebox.showerror("Hata", "İşletme adı boş olamaz.", parent=dlg); return
            set_config(self.con, "isletme_adi", ad)
            self._isletme = ad
            self.title(f"{ad} — Hesap Programı")
            dlg.destroy()

        ctk.CTkButton(dlg, text="Başlat", width=160, command=tamam).pack(pady=4)
        dlg.bind("<Return>", lambda _: tamam())

    def _ayarlar(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Ayarlar")
        dlg.geometry("360x180"); dlg.grab_set(); dlg.resizable(False, False)
        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        ctk.CTkLabel(frame, text="İşletme Adı:",
                     font=ctk.CTkFont(size=13)).grid(row=0, column=0, sticky="w", pady=8)
        e_ad = ctk.CTkEntry(frame, width=220)
        e_ad.insert(0, self._isletme)
        e_ad.grid(row=0, column=1, padx=10, pady=8)

        def kaydet():
            yeni = e_ad.get().strip()
            if not yeni:
                messagebox.showerror("Hata", "İşletme adı boş olamaz.", parent=dlg); return
            set_config(self.con, "isletme_adi", yeni)
            self._isletme = yeni
            self.title(f"{yeni} — Hesap Programı")
            dlg.destroy()
            messagebox.showinfo("Kaydedildi", f'İşletme adı "{yeni}" olarak güncellendi.')

        ctk.CTkButton(frame, text="Kaydet", command=kaydet).grid(
            row=1, column=0, columnspan=2, pady=14, sticky="ew")

    def _sessiz_kaydet(self):
        self.con.commit()

    def _hakkinda(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Hakkında")
        dlg.geometry("360x300"); dlg.grab_set(); dlg.resizable(False, False)
        ctk.CTkLabel(dlg, text="Esnaf Hesap Programı",
                     font=ctk.CTkFont(size=17, weight="bold")).pack(pady=(30,4))
        ctk.CTkLabel(dlg, text=f"Sürüm {VERSION}",
                     font=ctk.CTkFont(size=12), text_color="#555555").pack()
        ctk.CTkFrame(dlg, height=1, fg_color="#CCCCCC").pack(fill="x", padx=36, pady=18)
        ctk.CTkLabel(dlg, text=COPYRIGHT,
                     font=ctk.CTkFont(size=13, weight="bold")).pack()
        ctk.CTkLabel(dlg, text="Tüm hakları saklıdır.",
                     font=ctk.CTkFont(size=11), text_color="#666666").pack(pady=(2,8))
        ctk.CTkLabel(dlg, text="Bu yazılım izinsiz kopyalanamaz ve dağıtılamaz.",
                     font=ctk.CTkFont(size=10), text_color="#999999", wraplength=280).pack()
        ctk.CTkButton(dlg, text="Kapat", width=120,
                      command=dlg.destroy).pack(pady=22)

    def _yedek_al(self):
        now = datetime.today().strftime("%Y%m%d_%H%M")
        dest = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("Veritabanı Yedeği", "*.db"), ("Tümü", "*.*")],
            initialfile=f"hesap_yedek_{now}.db",
            title="Yedeği Kaydet"
        )
        if not dest:
            return
        self.con.commit()
        shutil.copy2(DB_PATH, dest)
        messagebox.showinfo("Yedek Alındı", f"Yedek başarıyla kaydedildi:\n{dest}")

    def _yedek_yukle(self):
        src = filedialog.askopenfilename(
            filetypes=[("Veritabanı Yedeği", "*.db"), ("Tümü", "*.*")],
            title="Geri Yüklenecek Yedeği Seç"
        )
        if not src:
            return
        if not messagebox.askyesno("Emin misiniz?",
                "Mevcut tüm veriler seçilen yedekle değiştirilecek.\nBu işlem geri alınamaz. Devam edilsin mi?"):
            return
        self.con.close()
        shutil.copy2(src, DB_PATH)
        self.con = sqlite3.connect(DB_PATH)
        self._show_placeholder()
        self.load_customers()
        messagebox.showinfo("Geri Yüklendi", "Yedek başarıyla geri yüklendi.")

    def _rapor_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        file_now = datetime.today().strftime("%Y%m%d_%H%M")
        dest = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile=f"hesap_raporu_{file_now}.xlsx",
            title="Raporu Kaydet"
        )
        if not dest:
            return

        wb = openpyxl.Workbook()

        def baslik_stili(ws, renk):
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill("solid", fgColor=renk)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 20

        # ── Sayfa 1: Borçlu Müşteriler ──
        ws1 = wb.active
        ws1.title = "Borçlu Müşteriler"
        ws1.append(["Müşteri Adı", "Telefon", "Borç (TL)"])
        baslik_stili(ws1, "2F5496")
        rows1 = self.con.cursor().execute("""
            SELECT m.musteri_adi, COALESCE(NULLIF(m.telefon,'nan'),''),
                   COALESCE((SELECT kalan FROM islemler WHERE musteri_id=m.id ORDER BY id DESC LIMIT 1),0) AS kalan
            FROM musteriler m WHERE m.arsiv=0
            ORDER BY kalan DESC, m.musteri_adi
        """).fetchall()
        toplam = 0.0
        for ad, tel, kalan in rows1:
            if kalan > 0:
                ws1.append([ad, tel, round(kalan, 2)])
                toplam += kalan
        ws1.append(["", "TOPLAM", round(toplam, 2)])
        for cell in ws1[ws1.max_row]:
            cell.font = Font(bold=True, size=11)
        for col, w in [("A", 32), ("B", 18), ("C", 16)]:
            ws1.column_dimensions[col].width = w

        # ── Sayfa 2: Günlük Satışlar ──
        ws2 = wb.create_sheet("Günlük Satışlar")
        ws2.append(["Tarih", "Ad Soyad", "Marka", "Ürün Türü", "Fiyat (TL)"])
        baslik_stili(ws2, "217346")
        rows2 = self.con.cursor().execute(
            "SELECT tarih, isim, marka, urun_turu, fiyat FROM gunluk_satis ORDER BY id DESC"
        ).fetchall()
        s_toplam = 0.0
        for r in rows2:
            ws2.append(list(r))
            s_toplam += r[4]
        ws2.append(["", "", "", "TOPLAM", round(s_toplam, 2)])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True, size=11)
        for col, w in [("A", 18), ("B", 28), ("C", 20), ("D", 20), ("E", 15)]:
            ws2.column_dimensions[col].width = w

        # ── Sayfa 3: Tüm İşlemler ──
        ws3 = wb.create_sheet("Müşteri İşlemleri")
        ws3.append(["Müşteri", "Tarih", "Açıklama", "Alışveriş (TL)", "Ödenen (TL)", "Kalan (TL)"])
        baslik_stili(ws3, "5B5B5B")
        rows3 = self.con.cursor().execute("""
            SELECT m.musteri_adi, i.tarih, i.aciklama, i.fiyat, i.odenen, i.kalan
            FROM islemler i JOIN musteriler m ON m.id=i.musteri_id
            ORDER BY m.musteri_adi, i.id
        """).fetchall()
        for r in rows3:
            ws3.append(list(r))
        for col, w in [("A", 28), ("B", 14), ("C", 26), ("D", 16), ("E", 16), ("F", 16)]:
            ws3.column_dimensions[col].width = w

        wb.save(dest)
        if messagebox.askyesno("Rapor Hazır", f"Rapor kaydedildi.\nExcel'de açmak ister misiniz?"):
            os.startfile(dest)

    def _kaydet_ve_cik(self):
        self.con.commit()
        self.con.close()
        self.after_cancel("all") if False else None
        self.quit()
        self.destroy()

    def _arsiv_toggle(self, mid, durum):
        self.con.cursor().execute("UPDATE musteriler SET arsiv=? WHERE id=?", (durum, mid))
        self.con.commit(); self._show_placeholder(); self.load_customers()

    def load_customers(self, *_):
        for item in self.tree.get_children(): self.tree.delete(item)
        q   = self.search_var.get().strip()
        fil = self.filter_var.get()
        sql = """
            SELECT m.id, m.musteri_adi,
                   COALESCE((SELECT kalan FROM islemler WHERE musteri_id=m.id ORDER BY id DESC LIMIT 1),0) AS kalan
            FROM musteriler m WHERE (m.musteri_adi LIKE ? OR m.telefon LIKE ?)
        """
        params = [f"%{q}%", f"%{q}%"]
        if fil == "Borçlular":
            sql += " AND COALESCE(m.arsiv,0)=0 AND kalan > 0"
        elif fil == "Tümü":
            sql += " AND COALESCE(m.arsiv,0)=0"
        elif fil == "Arşiv":
            sql += " AND COALESCE(m.arsiv,0)=1"
        sql += " ORDER BY m.musteri_adi COLLATE NOCASE"
        for row in self.con.cursor().execute(sql, params):
            mid, ad, kalan = row
            tag = "borclu" if kalan > 0 else "borcsuz"
            lbl = f"{kalan:,.2f} TL" if kalan != 0 else "—"
            self.tree.insert("", "end", iid=str(mid), values=(ad, lbl), tags=(tag,))

    def on_select(self, _):
        sel = self.tree.selection()
        if sel: self._show_customer(int(sel[0]))

    def _show_customer(self, mid):
        for w in self.right.winfo_children(): w.destroy()
        cur = self.con.cursor()
        musteri = cur.execute("SELECT musteri_adi,telefon FROM musteriler WHERE id=?",(mid,)).fetchone()
        if not musteri: return
        ad, tel = musteri
        kalan = cur.execute(
            "SELECT COALESCE(kalan,0) FROM islemler WHERE musteri_id=? ORDER BY id DESC LIMIT 1",(mid,)
        ).fetchone()
        kalan = kalan[0] if kalan else 0.0

        header = ctk.CTkFrame(self.right, fg_color="#2F5496", corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text=ad, font=ctk.CTkFont(size=17, weight="bold"),
                     text_color="white").pack(side="left", padx=16, pady=10)
        tel_lbl = f"📞 {tel}" if tel and tel not in ("nan","") else ""
        ctk.CTkLabel(header, text=tel_lbl, text_color="#BDD7EE").pack(side="left", padx=4)
        color = "#C00000" if kalan > 0 else "#217346"
        ctk.CTkLabel(header, text=f"Bakiye: {kalan:,.2f} TL",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=color, fg_color="white" if kalan > 0 else "transparent",
                     corner_radius=6).pack(side="right", padx=16)

        tbl_frame = ctk.CTkFrame(self.right, fg_color="transparent")
        tbl_frame.pack(fill="both", expand=True, padx=12, pady=8)
        det = ttk.Treeview(tbl_frame,
            columns=("tarih","aciklama","fiyat","odenen","kalan"),
            show="headings", style="Custom.Treeview", selectmode="browse")
        for c,lbl,w,anc in [("tarih","Tarih",90,"center"),("aciklama","Açıklama",200,"w"),
                              ("fiyat","Alışveriş",100,"e"),("odenen","Ödenen",100,"e"),("kalan","Kalan",100,"e")]:
            det.heading(c,text=lbl); det.column(c,width=w,anchor=anc)
        for r in cur.execute(
            "SELECT tarih,aciklama,fiyat,odenen,kalan FROM islemler WHERE musteri_id=? ORDER BY id",(mid,)
        ).fetchall():
            det.insert("","end",values=(r[0],r[1],
                f"{r[2]:,.2f}" if r[2] else "",
                f"{r[3]:,.2f}" if r[3] else "",
                f"{r[4]:,.2f}" if r[4] is not None else ""))
        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=det.yview)
        det.configure(yscrollcommand=vsb.set)
        det.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        btn_bar = ctk.CTkFrame(self.right, fg_color="transparent")
        btn_bar.pack(fill="x", padx=12, pady=(0,10))
        ctk.CTkButton(btn_bar, text="+ Alışveriş Ekle", width=150, fg_color="#2F5496",
                      command=lambda: self._add_islem(mid,"alis")).pack(side="left", padx=3)
        ctk.CTkButton(btn_bar, text="+ Ödeme Ekle", width=150, fg_color="#217346",
                      command=lambda: self._add_islem(mid,"odeme")).pack(side="left", padx=3)
        ctk.CTkButton(btn_bar, text="✏ Düzenle", width=120, fg_color="#5B5B5B",
                      command=lambda: self._edit_customer(mid,ad,tel)).pack(side="left", padx=3)
        arsiv_durum = cur.execute("SELECT COALESCE(arsiv,0) FROM musteriler WHERE id=?",(mid,)).fetchone()[0]
        if arsiv_durum:
            ctk.CTkButton(btn_bar, text="♻ Arşivden Çıkar", width=140, fg_color="#217346",
                          command=lambda: self._arsiv_toggle(mid,0)).pack(side="left", padx=3)
        else:
            ctk.CTkButton(btn_bar, text="📦 Arşivle", width=110, fg_color="#ED7D31",
                          command=lambda: self._arsiv_toggle(mid,1)).pack(side="left", padx=3)
        ctk.CTkButton(btn_bar, text="🗑 Sil", width=90, fg_color="#C00000",
                      command=lambda: self._delete_customer(mid)).pack(side="right", padx=3)

    def _add_islem(self, mid, tip):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Alışveriş Ekle" if tip=="alis" else "Ödeme Ekle")
        dlg.geometry("400x280"); dlg.grab_set(); dlg.resizable(False,False)
        cur = self.con.cursor()
        son_kalan = cur.execute(
            "SELECT COALESCE(kalan,0) FROM islemler WHERE musteri_id=? ORDER BY id DESC LIMIT 1",(mid,)
        ).fetchone()
        son_kalan = son_kalan[0] if son_kalan else 0.0
        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=16)
        ctk.CTkLabel(frame, text="Tarih:").grid(row=0, column=0, sticky="w", pady=6)
        e_tarih = ctk.CTkEntry(frame, width=210)
        e_tarih.insert(0, datetime.today().strftime("%d.%m.%Y %H:%M"))
        e_tarih.grid(row=0, column=1, pady=6, padx=8)
        ctk.CTkLabel(frame, text="Açıklama:").grid(row=1, column=0, sticky="w", pady=6)
        e_acikl = ctk.CTkEntry(frame, width=210)
        if tip == "odeme": e_acikl.insert(0, "Borç Ödeme")
        e_acikl.grid(row=1, column=1, pady=6, padx=8)
        lbl = "Tutar (TL):" if tip=="alis" else "Ödenen (TL):"
        ctk.CTkLabel(frame, text=lbl).grid(row=2, column=0, sticky="w", pady=6)
        e_tutar = ctk.CTkEntry(frame, width=210)
        e_tutar.grid(row=2, column=1, pady=6, padx=8)

        def kaydet():
            try: tutar = float(e_tutar.get().replace(",","."))
            except ValueError:
                messagebox.showerror("Hata","Geçerli bir tutar girin.",parent=dlg); return
            fiyat      = tutar if tip=="alis" else 0.0
            yeni_kalan = son_kalan + tutar if tip=="alis" else son_kalan - tutar
            cur.execute(
                "INSERT INTO islemler (musteri_id,tarih,aciklama,fiyat,odenen,kalan) VALUES (?,?,?,?,?,?)",
                (mid, e_tarih.get().strip(), e_acikl.get().strip(),
                 fiyat, tutar if tip=="odeme" else 0.0, round(yeni_kalan,2))
            )
            self.con.commit(); dlg.destroy()
            self._show_customer(mid); self.load_customers()

        ctk.CTkButton(frame, text="Kaydet",
                      fg_color="#2F5496" if tip=="alis" else "#217346",
                      command=kaydet).grid(row=3, column=0, columnspan=2, pady=14, sticky="ew")

    def new_customer(self): self._customer_form(None,"","")
    def _edit_customer(self,mid,ad,tel): self._customer_form(mid,ad,tel)

    def _customer_form(self, mid, ad, tel):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Yeni Müşteri" if mid is None else "Müşteri Düzenle")
        dlg.geometry("340x210"); dlg.grab_set(); dlg.resizable(False,False)
        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=16)
        ctk.CTkLabel(frame, text="Müşteri Adı:").grid(row=0, column=0, sticky="w", pady=6)
        e_ad = ctk.CTkEntry(frame, width=200); e_ad.insert(0, ad)
        e_ad.grid(row=0, column=1, padx=8)
        ctk.CTkLabel(frame, text="Telefon:").grid(row=1, column=0, sticky="w", pady=6)
        e_tel = ctk.CTkEntry(frame, width=200)
        e_tel.insert(0, tel if tel not in ("nan","") else "")
        e_tel.grid(row=1, column=1, padx=8)

        def kaydet():
            ad_v = e_ad.get().strip()
            if not ad_v:
                messagebox.showerror("Hata","Müşteri adı boş olamaz.",parent=dlg); return
            cur = self.con.cursor()
            if mid is None:
                cur.execute("INSERT INTO musteriler (musteri_adi,telefon) VALUES (?,?)", (ad_v, e_tel.get().strip()))
                new_id = cur.lastrowid
            else:
                cur.execute("UPDATE musteriler SET musteri_adi=?,telefon=? WHERE id=?",
                            (ad_v, e_tel.get().strip(), mid))
                new_id = mid
            self.con.commit(); dlg.destroy()
            self.load_customers(); self._show_customer(new_id)

        ctk.CTkButton(frame, text="Kaydet", command=kaydet).grid(
            row=2, column=0, columnspan=2, pady=14, sticky="ew")

    def _delete_customer(self, mid):
        cur = self.con.cursor()
        ad = cur.execute("SELECT musteri_adi FROM musteriler WHERE id=?",(mid,)).fetchone()[0]
        if not messagebox.askyesno("Onay",f'"{ad}" silinsin mi? Tüm işlemleri de silinir.',parent=self): return
        cur.execute("DELETE FROM islemler   WHERE musteri_id=?",(mid,))
        cur.execute("DELETE FROM musteriler WHERE id=?",        (mid,))
        self.con.commit(); self._show_placeholder(); self.load_customers()


# ── BAŞLANGIÇ ─────────────────────────────────────────────────
if __name__ == "__main__":
    show_splash()
    con = init_db()
    app = App(con)
    app.mainloop()
    con.close()
